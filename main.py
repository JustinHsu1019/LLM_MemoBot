from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from datetime import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import io, time, openpyxl
from openpyxl.styles import Side, Border

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--disable-blink-features=BlockCredentialedSubresources")
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.set_page_load_timeout(60)

def get_consecutive_days(driver, stock, retries=3):
    url = f"https://www.cmoney.tw/finance/{stock}/f00036"
    attempt = 0
    while attempt < retries:
        try:
            driver.get(url)
            driver.implicitly_wait(10)
            table = driver.find_element(By.CSS_SELECTOR, "table.tb.tb1")
            rows = table.find_elements(By.TAG_NAME, "tr")
            data = []

            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                if cols:
                    data.append([col.text.strip() for col in cols])

            columns = ["日期", "外資買賣超", "投信買賣超", "自營商買賣超", "三大法人合計", "外資-持股張數", "外資-持股比率", "投信-持股張數", "投信-持股比率", "自營商-持股張數", "自營商-持股比率"]
            df = pd.DataFrame(data, columns=columns)

            df["投信買賣超"] = pd.to_numeric(df["投信買賣超"].str.replace(',', ''), errors='coerce')

            def get_consecutive_buy_sell(column):
                if df[column].empty:
                    return "No data"
                count = 0
                sign = 1 if df[column].iloc[0] > 0 else -1
                for value in df[column]:
                    if (sign > 0 and value > 0) or (sign < 0 and value < 0):
                        count += 1
                    else:
                        break
                return "連續" if count == len(df[column]) else count

            return get_consecutive_buy_sell("投信買賣超")
        except Exception as e:
            print(f"Error processing stock {stock}, attempt {attempt + 1}: {e}")
            attempt += 1
            time.sleep(5)
    return "Error"

def scrape_data(driver, url):
    driver.get(url)
    driver.implicitly_wait(10)
    table = driver.find_element(By.CSS_SELECTOR, "table.tb.tb1")
    rows = table.find_elements(By.TAG_NAME, "tr")
    data = []

    for row in rows:
        cols = row.find_elements(By.TAG_NAME, "td")
        if cols:
            data.append([col.text.strip() for col in cols])
    
    return data

url1 = "https://www.cmoney.tw/finance/f00065.aspx?t=0&o=1&o2=3&d=1"
data1 = scrape_data(driver, url1)
url2 = "https://www.cmoney.tw/finance/f00065.aspx?t=0&o=2&o2=3&d=1"
data2 = scrape_data(driver, url2)

columns = ["投信買 / 賣超排名", "代碼 / 股票名稱", "外資", "投信買超金額", "自營商", "合計"]
adjusted_columns = ["投信買 / 賣超排名", "代碼 / 股票名稱", "投信買超金額", "外資"]

df1 = pd.DataFrame(data1, columns=columns)
df2 = pd.DataFrame(data2[:5], columns=columns)
df1 = df1[adjusted_columns]
df2 = df2[adjusted_columns]
blank_row = pd.DataFrame([[""] * len(adjusted_columns)], columns=adjusted_columns)
combined_data = pd.concat([df1, blank_row, df2], ignore_index=True)
temp_num = combined_data["代碼 / 股票名稱"].apply(lambda x: x.split(" ")[0] if " " in x else "")

def process_stock_code(stock_code):
    if stock_code.strip():
        return get_consecutive_days(driver, stock_code)
    return ""

combined_data["連續買 / 賣超"] = temp_num.apply(process_stock_code)

driver.quit()

def upload_to_drive(file_content):
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    SERVICE_ACCOUNT_FILE = 'cred.json'

    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    drive_service = build('drive', 'v3', credentials=credentials)

    file_metadata = {
        'name': 'data.xlsx',
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    media = MediaIoBaseUpload(file_content, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    file_id = file.get('id')

    drive_service.permissions().create(
        fileId=file_id,
        body={'type': 'anyone', 'role': 'reader'}
    ).execute()

    return f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"

output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
    combined_data.to_excel(writer, sheet_name='工作表', startrow=2, startcol=1, index=False)
    workbook = writer.book
    worksheet = writer.sheets['工作表']
    
    current_date = datetime.now().strftime("%Y/%m/%d")
    worksheet.cell(row=2, column=2).value = current_date
    worksheet.cell(row=2, column=2).font = openpyxl.styles.Font(name='Microsoft JhengHei', size=12)

    red_fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    green_fill = openpyxl.styles.PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    center_align = openpyxl.styles.Alignment(horizontal='center')
    left_align = openpyxl.styles.Alignment(horizontal='left')
    right_align = openpyxl.styles.Alignment(horizontal='right')
    white_font = openpyxl.styles.Font(name='Microsoft JhengHei', color='FFFFFF', bold=True, size=12)

    for row in worksheet.iter_rows(min_row=4, max_row=33, min_col=2, max_col=2):
        for cell in row:
            cell.fill = red_fill
            cell.font = white_font

    for row in worksheet.iter_rows(min_row=35, max_row=39, min_col=2, max_col=2):
        for cell in row:
            cell.fill = green_fill
            cell.font = white_font

    for col_num in range(2, worksheet.max_column + 1):
        for row_num in range(3, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            if col_num == 4 or col_num == 5:
                cell.alignment = right_align
            elif col_num == 3:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            cell.font = white_font if cell.fill == red_fill or cell.fill == green_fill else openpyxl.styles.Font(name='Microsoft JhengHei', size=12)

    header_font = openpyxl.styles.Font(name='Microsoft JhengHei', bold=True, size=12)
    for cell in worksheet[3]:
        cell.font = header_font
        cell.alignment = center_align
        new_border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style=None),
                                            right=openpyxl.styles.Side(border_style=None),
                                            top=openpyxl.styles.Side(border_style=None),
                                            bottom=openpyxl.styles.Side(border_style=None))
        cell.border = new_border

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 10)
        worksheet.column_dimensions[column].width = adjusted_width

    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 20

    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )

    for row in range(2, 40):
        worksheet.cell(row=row, column=2).border = Border(left=Side(style='thick', color='000000'))
        worksheet.cell(row=row, column=7).border = Border(right=Side(style='thick', color='000000'))

    for col in range(2, 8):
        worksheet.cell(row=2, column=col).border = Border(top=Side(style='thick', color='000000'))
        worksheet.cell(row=39, column=col).border = Border(bottom=Side(style='thick', color='000000'))

    for row in range(2, 40):
        worksheet.cell(row=row, column=2).border = Border(
            left=Side(style='thick', color='000000'),
            bottom=worksheet.cell(row=row, column=2).border.bottom,
            right=worksheet.cell(row=row, column=2).border.right,
            top=worksheet.cell(row=row, column=2).border.top
        )
        worksheet.cell(row=row, column=7).border = Border(
            right=Side(style='thick', color='000000'),
            bottom=worksheet.cell(row=row, column=7).border.bottom,
            left=worksheet.cell(row=row, column=7).border.left,
            top=worksheet.cell(row=row, column=7).border.top
        )

output.seek(0)
drive_link = upload_to_drive(output)

def update_google_sheet(date, link):
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    SERVICE_ACCOUNT_FILE = 'cred.json'

    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    sheets_service = build('sheets', 'v4', credentials=credentials)
    spreadsheet_id = '1ZNCRPM9N6qZBSPvP_ggou91xHjWxfM53MjfvlXEaweg'
    range_name = 'A:B'
    
    body = {
        'values': [
            [date, link]
        ]
    }

    result = sheets_service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id, range=range_name,
        valueInputOption='RAW', body=body).execute()

update_google_sheet(current_date, drive_link)
