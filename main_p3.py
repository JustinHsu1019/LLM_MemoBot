from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import io
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.set_page_load_timeout(60)

def scrape_ranking_data(driver, url, top_n):
    driver.get(url)
    driver.implicitly_wait(10)
    table = driver.find_element(By.CSS_SELECTOR, "table.tb1")
    rows = table.find_elements(By.TAG_NAME, "tr")
    data = []

    for index, row in enumerate(rows):
        if index >= top_n:
            break
        cols = row.find_elements(By.TAG_NAME, "td")
        if cols:
            rank = cols[0].text.strip()
            stock_name = cols[1].text.strip()
            stock_code = stock_name.split(" ")[0]
            stock_ratio = cols[2].text.strip()
            day_trade_ratio = cols[3].text.strip()
            data.append([rank, stock_name, stock_code, stock_ratio, day_trade_ratio])

    columns = ["券資比排名", "代碼 / 股票名稱", "股票代碼", "券資比", "當沖比率"]
    return pd.DataFrame(data, columns=columns)


def calculate_consecutive_days(driver, stock_code):
    url = f"https://www.cmoney.tw/finance/{stock_code}/f00037"
    driver.get(url)
    driver.implicitly_wait(10)

    table = driver.find_element(By.CSS_SELECTOR, "table.tb.tb1")
    rows = table.find_elements(By.TAG_NAME, "tr")
    
    stock_ratios = []

    for row in rows:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) > 8:
            ratio_text = cols[7].text.strip().replace(',', '')
            try:
                stock_ratios.append(float(ratio_text))
            except ValueError:
                stock_ratios.append(None)
    
    print("股票券資比數據: ", stock_ratios)
    
    return get_consecutive_change(stock_ratios)

def get_consecutive_change(stock_ratios):
    if not stock_ratios or len(stock_ratios) < 2:
        print("輸入資料無效或天數不足")
        return 0

    count = 1
    sign = 0

    print(f"初始資料: {stock_ratios}")

    for i in range(len(stock_ratios) - 1):
        current_sign = 1 if stock_ratios[i + 1] > stock_ratios[i] else -1

        print(f"比較: {stock_ratios[i]} 和 {stock_ratios[i + 1]} -> 當前趨勢: {'遞增' if current_sign == 1 else '遞減'}")

        if sign == 0:
            sign = current_sign
            print(f"設定初始趨勢為: {'遞增' if sign == 1 else '遞減'}")

        if current_sign == sign:
            count += 1
            print(f"當前趨勢一致，連續天數增加至: {count}")
        else:
            print(f"趨勢改變，最終連續天數為: {count}")
            count = count - 1
            return -count if sign == 1 else count

    print(f"趨勢沒有改變，最終連續天數為: {count}")
    count = count - 1
    return -count if sign == 1 else count

def upload_to_drive(file_content):
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    SERVICE_ACCOUNT_FILE = 'cred.json'

    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    drive_service = build('drive', 'v3', credentials=credentials)

    file_metadata = {
        'name': 'stock_report.xlsx',
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    media = MediaIoBaseUpload(file_content, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    file_id = file.get('id')

    drive_service.permissions().create(
        fileId=file_id,
        body={'type': 'anyone', 'role': 'reader'}
    ).execute()

    return f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"

def update_google_sheet(date, link):
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    SERVICE_ACCOUNT_FILE = 'cred.json'

    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    sheets_service = build('sheets', 'v4', credentials=credentials)
    spreadsheet_id = '1XgZF7I9HyjRveGid483HFLvmrup8CgUdJhNWbiZDFrs'
    range_name = 'A:B'
    
    body = {
        'values': [
            [date, link]
        ]
    }

    result = sheets_service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id, range=range_name,
        valueInputOption='RAW', body=body).execute()

url_ranking = "https://www.cmoney.tw/finance/f00068.aspx"
top_n = 11

ranking_data = scrape_ranking_data(driver, url_ranking, top_n)

ranking_data["券資比遞增遞減"] = ranking_data["股票代碼"].apply(lambda stock_code: calculate_consecutive_days(driver, stock_code))
ranking_data.drop(columns=["股票代碼"], inplace=True)

cols = ["券資比排名", "代碼 / 股票名稱", "券資比", "券資比遞增遞減", "當沖比率"]
ranking_data = ranking_data[cols]

driver.quit()

output = io.BytesIO()
today_date = datetime.now().strftime("%Y/%m/%d")

with pd.ExcelWriter(output, engine='openpyxl') as writer:
    ranking_data.to_excel(writer, sheet_name='工作表', index=False, startrow=2, startcol=1)
    workbook = writer.book
    worksheet = writer.sheets['工作表']

    font = Font(name='Microsoft JhengHei', size=12)
    bold_font = Font(name='Microsoft JhengHei', size=12, bold=True)
    red_font = Font(name='Microsoft JhengHei', color='FFFFFF', bold=True, size=12)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    center_align = Alignment(horizontal='center')

    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )

    for col in worksheet.columns:
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = 20

    for cell in worksheet[3]:
        cell.font = bold_font
        cell.alignment = center_align

    for row in worksheet.iter_rows(min_row=4):
        for cell in row:
            cell.font = font
            cell.alignment = center_align

    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
        for cell in row:
            cell.font = red_font
            cell.fill = red_fill

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row in range(2, max_row + 2):
        if worksheet.cell(row=row, column=2).value is not None:
            worksheet.cell(row=row, column=2).border = Border(left=Side(style='thick', color='000000'))
            worksheet.cell(row=row, column=max_col).border = Border(right=Side(style='thick', color='000000'))

    for col in range(2, max_col + 1):
        worksheet.cell(row=2, column=col).border = Border(top=Side(style='thick', color='000000'))
        worksheet.cell(row=max_row, column=col).border = Border(bottom=Side(style='thick', color='000000'))

    worksheet.cell(row=2, column=2).border = Border(left=Side(style='thick', color='000000'), top=Side(style='thick', color='000000'))
    worksheet.cell(row=2, column=max_col).border = Border(right=Side(style='thick', color='000000'), top=Side(style='thick', color='000000'))
    worksheet.cell(row=max_row, column=2).border = Border(left=Side(style='thick', color='000000'), bottom=Side(style='thick', color='000000'))
    worksheet.cell(row=max_row, column=max_col).border = Border(right=Side(style='thick', color='000000'), bottom=Side(style='thick', color='000000'))

    date_cell = worksheet.cell(row=2, column=2)
    date_cell.value = today_date
    date_cell.font = bold_font
    date_cell.alignment = center_align
    
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    
    worksheet.cell(row=3, column=3).border = no_border
    worksheet.cell(row=3, column=4).border = no_border
    worksheet.cell(row=3, column=5).border = no_border

output.seek(0)

drive_link = upload_to_drive(output)

update_google_sheet(today_date, drive_link)

print("Report generated and uploaded successfully.")
